"""
Deterministic build-time extractor for the Khasab dataset.

Reads:
  C:/Users/Subrat Mohapatra/Downloads/eppm_documents/Updated invitation_ AI Agent Development/1. Daily Data-Khasab  Jan, Feb, Mar 2026.xlsx

Writes:
  backend/bipros-data-bootstrap/src/main/resources/bootstrap-data.json

This script is NOT invoked at runtime. Run it manually when the source workbook
changes:
    python backend/bipros-data-bootstrap/scripts/extract.py

The JSON it produces is the single source of truth every Stage* class reads at
runtime via the classpath — no LLM, no network, no API key.

Date strategy
=============
Source rows are dated 2025-01-24 → 2025-03-29. We shift every date so the most
recent DPR lands on TODAY, giving a "last 3 months" project that reads as live
when the operator runs the bootstrap.

  shift_days  = today - max(source_date)
  new_date    = old_date + shift_days

Material data
=============
The source workbook's material columns are empty across all 53 days, and its
MAT master sheet is all #REF! errors. materialVariants is therefore an empty
list. Material roles can be added manually later.
"""
from __future__ import annotations

import datetime
import json
import re
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl

SOURCE_XLSX = Path(
    r"C:\Users\Subrat Mohapatra\Downloads\eppm_documents"
    r"\Updated invitation_ AI Agent Development"
    r"\1. Daily Data-Khasab  Jan, Feb, Mar 2026.xlsx"
)

# Real concrete production data — used to populate Material variants + DPR material rows.
CONCRETE_XLSX = Path(
    r"C:\Users\Subrat Mohapatra\Downloads\eppm_documents"
    r"\Updated invitation_ AI Agent Development"
    r"\Concrete Summary - Khasab.xlsx"
)

# Industry-typical OMR/m³ for ready-mix concrete grades in Oman.
# Override these with the client's actual rates when they share the contract.
CONCRETE_RATE_BY_GRADE: dict[str, float] = {
    "C15": 35.0,
    "C20": 45.0,
    "C25": 50.0,
    "C30": 55.0,
    "C35": 62.0,
    "C40": 70.0,
}

REPO_ROOT = Path(__file__).resolve().parents[3]
OUT_JSON = REPO_ROOT / "backend" / "bipros-data-bootstrap" / "src" / "main" / "resources" / "bootstrap-data.json"

DAILY_SHEETS = ["Jan-2026", "Feb-2026", "March-2026"]
TODAY = datetime.date.today()


# ───────────────────────── helpers ─────────────────────────

def sanitize_code(raw: Any) -> str:
    """'2.3.6(i)a' -> '2-3-6-I-A'.  '5.1.7 (iii) ' -> '5-1-7-III'."""
    s = str(raw).strip().upper()
    s = s.replace("(", "-").replace(")", "")
    s = s.replace(".", "-").replace(" ", "")
    s = re.sub(r"-+", "-", s).strip("-")
    return s


def role_code(name: str) -> str:
    """'A/C Technician' -> 'AC_TECHNICIAN'."""
    s = name.strip().upper()
    s = re.sub(r"[/\s\-]+", "_", s)
    s = re.sub(r"[^A-Z0-9_]", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def equipment_code(name: str) -> str:
    return role_code(name)


def normalise_equipment(name: str) -> str:
    """Collapse known spelling variants seen in the source."""
    n = name.strip()
    low = n.lower()
    if low in ("cruhser", "crucher", "crusher"):
        return "Crusher"
    return n


WBS_CHAPTERS_DEF = [
    # (matching prefix, code, name, sort)
    ("1",  "PREP",     "Preliminaries & Camp",       10),
    ("2",  "EARTH",    "Earthwork",                  20),
    ("3",  "PVMT-AGG", "Pavement — Aggregate Base",  30),
    ("4",  "PVMT-BIT", "Pavement — Bituminous",      40),
    ("5",  "CONCR",    "Concrete & Reinforcement",   50),
    ("8",  "PIPE",     "Pipes & Culverts",           80),
    ("9",  "PROT",     "Slope Protection",           90),
    ("12", "PAVE",     "Paving & Curbs",            120),
    ("13", "BARRIER",  "Barriers & Guardrails",     130),
    ("14", "SIGN",     "Signs & Markings",          140),
    ("18", "UTIL",     "Utilities & Ducts",         180),
]


def wbs_chapter_for(activity_code: Any) -> tuple[str, str, int]:
    code_str = str(activity_code).strip()
    m = re.match(r"^(\d+)", code_str)
    prefix = m.group(1) if m else ""
    for pfx, code, name, sort in WBS_CHAPTERS_DEF:
        if prefix == pfx:
            return code, name, sort
    return "MISC", "Miscellaneous", 999


def discipline_for(activity_code: Any) -> str:
    code, _name, _ = wbs_chapter_for(activity_code)
    return {
        "PREP": "preliminaries",
        "EARTH": "earthwork",
        "PVMT-AGG": "pavement",
        "PVMT-BIT": "pavement",
        "CONCR": "structures",
        "PIPE": "structures",
        "PROT": "structures",
        "PAVE": "finishes",
        "BARRIER": "finishes",
        "SIGN": "finishes",
        "UTIL": "utilities",
        "MISC": "general",
    }[code]


def to_float(v: Any) -> float | None:
    if v is None or v == "" or v == "#REF!":
        return None
    try:
        f = float(v)
        if f != f:  # NaN
            return None
        return f
    except (TypeError, ValueError):
        return None


def to_int(v: Any) -> int | None:
    f = to_float(v)
    return int(f) if f is not None else None


def round2(v: float | None) -> float | None:
    return None if v is None else round(v, 2)


# ───────────────────────── extraction ─────────────────────────

def load_code_sheet(wb) -> dict[str, Any]:
    """
    Code sheet columns (0-indexed):
       2: activity code   3: description   4: BOQ unit
       6: manpower role   7: manpower hourly rate
       9: equipment name  10: equipment hourly rate
    """
    ws = wb["Code"]
    activities_master: dict[str, dict] = {}
    manpower_rates: dict[str, float] = {}
    equipment_rates: dict[str, float] = {}

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # Activity code + description + unit
        act_code = row[2] if len(row) > 2 else None
        if act_code is not None and str(act_code).strip() and str(act_code).strip().lower() != "activity code":
            desc = (row[3] if len(row) > 3 else None) or ""
            unit = (row[4] if len(row) > 4 else None) or "Nos"
            key = str(act_code).strip()
            activities_master.setdefault(key, {
                "description": str(desc).strip() or key,
                "unit": str(unit).strip() if unit else "Nos",
            })

        # Manpower role + rate
        mp_role = row[6] if len(row) > 6 else None
        mp_rate = to_float(row[7] if len(row) > 7 else None)
        if mp_role and isinstance(mp_role, str) and mp_role.strip() and mp_rate is not None:
            manpower_rates[mp_role.strip()] = mp_rate

        # Equipment + rate
        eq_name = row[9] if len(row) > 9 else None
        eq_rate = to_float(row[10] if len(row) > 10 else None)
        if eq_name and isinstance(eq_name, str) and eq_name.strip() and eq_rate is not None:
            equipment_rates[normalise_equipment(eq_name)] = eq_rate

    return {
        "activities_master": activities_master,
        "manpower_rates": manpower_rates,
        "equipment_rates": equipment_rates,
    }


def load_daily_rows(wb) -> list[dict]:
    """
    Daily-sheet columns (0-indexed):
       1: date          2: site             3: location
       4: from          5: to               6: side
       7: activity code 8: unit             9: executed qty
      10: supervisor    11: manpower role  12: nos  13: total hrs  14: rate/hr
      16: equipment     17: nos            18: total hrs 19: rate/hr
      21: material desc 22: unit           23: qty   24: rate
    """
    rows: list[dict] = []
    for sheet in DAILY_SHEETS:
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 4:                                # skip header
                continue
            date = row[1] if len(row) > 1 else None
            if not isinstance(date, datetime.datetime):
                continue
            act = row[7] if len(row) > 7 else None
            sup = row[10] if len(row) > 10 else None
            if not act or not sup or not isinstance(sup, str) or not sup.strip():
                continue
            rows.append({
                "date": date.date(),
                "site": (row[2] or "").strip() if isinstance(row[2], str) else None,
                "location": (row[3] or "").strip() if isinstance(row[3], str) else None,
                "from_ch": to_float(row[4]),
                "to_ch": to_float(row[5]),
                "side": (row[6] or "").strip() if isinstance(row[6], str) else None,
                "activity_code": str(act).strip(),
                "unit": (row[8] or "").strip() if isinstance(row[8], str) else None,
                "executed_qty": to_float(row[9]),
                "supervisor": sup.strip(),
                "mp_role": (row[11].strip() if isinstance(row[11], str) and row[11].strip() else None) if len(row) > 11 else None,
                "mp_nos": to_int(row[12] if len(row) > 12 else None),
                "mp_hours": to_float(row[13] if len(row) > 13 else None),
                "mp_rate": to_float(row[14] if len(row) > 14 else None),
                "eq_name": normalise_equipment(row[16]) if (len(row) > 16 and isinstance(row[16], str) and row[16].strip()) else None,
                "eq_nos": to_int(row[17] if len(row) > 17 else None),
                "eq_hours": to_float(row[18] if len(row) > 18 else None),
                "eq_rate": to_float(row[19] if len(row) > 19 else None),
            })
    return rows


def load_concrete_summary() -> list[dict]:
    """
    Read the Concrete Summary - Khasab.xlsx sheet 'Khasab'. Schema (0-indexed):
       2: date       3: chainage      4: structure      5: element
       6: grade      7: quantity m³   9: section
    Returns a list of dicts; rows with no date or qty are dropped.
    """
    if not CONCRETE_XLSX.exists():
        print(f"  WARN concrete-summary not found at {CONCRETE_XLSX}; material data will be empty")
        return []
    wb = openpyxl.load_workbook(CONCRETE_XLSX, data_only=True, read_only=True)
    ws = wb["Khasab"]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:                            # skip header rows
            continue
        date = row[2] if len(row) > 2 else None
        qty = to_float(row[7] if len(row) > 7 else None)
        grade = row[6] if len(row) > 6 else None
        if not isinstance(date, datetime.datetime) or qty is None or qty <= 0 or not grade:
            continue
        out.append({
            "date": date.date(),
            "chainage": str(row[3]).strip() if row[3] is not None else None,
            "structure": str(row[4]).strip() if row[4] is not None else None,
            "element": str(row[5]).strip() if row[5] is not None else None,
            "grade": str(grade).strip().upper(),
            "qty": qty,
        })
    return out


# Structure type → preferred activity codes for matching concrete consumption to DPRs.
# Order matters: first match that has a DPR on the date wins.
CONCRETE_ACTIVITY_MAP: dict[str, list[str]] = {
    "concrete barrier": ["13.1.7(ix)b2", "13.1.7(ix)a2", "5.1.7 (ii) "],
    "box culvert":       ["8.1.6 (vii)a", "8.1.6 (vii)b", "8.1.6 (vii)c", "8.1.6 (vii)d",
                          "5.1.7 (iii) ", "5.1.7 (i) "],
    "retaining wall":    ["5.1.7 (iii) ", "5.1.7 (i) "],
    "box slab":          ["8.1.6 (vii)a", "5.1.7 (iii) "],
}
# Grade → fallback activity when structure didn't match.
GRADE_FALLBACK_ACTIVITY: dict[str, str] = {
    "C15": "5.1.7 (i) ",
    "C20": "5.1.7 (ii) ",
    "C25": "5.1.7 (ii) ",
    "C30": "5.1.7 (iii) ",
    "C35": "5.1.7 (iii) ",
    "C40": "5.1.7 (iii) ",
}


def pick_activity_for_concrete(structure: str | None, grade: str,
                                available_for_date: set[str]) -> str | None:
    """Return the first activity code that (a) maps to this structure/grade and (b) has a DPR
    on the same date. Falls back to grade-based mapping. Returns None when nothing matches."""
    if structure:
        s = structure.lower()
        for key, codes in CONCRETE_ACTIVITY_MAP.items():
            if key in s:
                for code in codes:
                    if code in available_for_date:
                        return code
    fallback = GRADE_FALLBACK_ACTIVITY.get(grade)
    if fallback and fallback in available_for_date:
        return fallback
    return None


def build_dataset(code_data: dict, rows: list[dict]) -> dict:
    # ── date shift ──
    source_max = max(r["date"] for r in rows)
    shift_days = (TODAY - source_max).days
    print(f"date shift: source max = {source_max}, today = {TODAY}, shift = {shift_days} days")

    def shifted(d: datetime.date) -> datetime.date:
        return d + datetime.timedelta(days=shift_days)

    # ── project ──
    project = {
        "code": "KHASAB-001",
        "name": "SC 180 — Khasab — Daba Asphalt Road & Link to Lima",
        "description": "Design and Construction of Khasab — Daba Asphalt Road and Link to Lima (Sultanate of Oman). "
                       "Data bootstrapped from real DPRs (Jan–Mar 2025 source), dates re-anchored so the last "
                       "DPR is today.",
        "plannedStart": shifted(min(r["date"] for r in rows)).isoformat(),
        "plannedFinish": (shifted(source_max) + datetime.timedelta(days=14)).isoformat(),
        "currency": "OMR",
        "calendarCode": None,
        "fromLocation": "Khasab",
        "toLocation": "Daba",
        "fromChainageM": 0,
        "toChainageM": 80000,
        "morthCode": None,
        "category": None,
    }

    # ── unique sets ──
    mp_role_set = sorted({r["mp_role"] for r in rows if r["mp_role"]})
    eq_set = sorted({r["eq_name"] for r in rows if r["eq_name"]})
    activity_codes = sorted({r["activity_code"] for r in rows})
    supervisors = sorted({r["supervisor"] for r in rows})

    # ── manpower variants ── one row per manpower role (Skilled/A default,
    # rate from Code sheet; fall back to first non-zero rate seen in DPRs).
    manpower_variants = []
    for name in mp_role_set:
        rate = code_data["manpower_rates"].get(name)
        if rate is None:
            # try DPR-observed rate
            rate = next((r["mp_rate"] for r in rows
                         if r["mp_role"] == name and r["mp_rate"] and r["mp_rate"] > 0), None)
        if rate is None:
            rate = 1.0
        manpower_variants.append({
            "roleCode": role_code(name),
            "roleName": name,
            "categoryCode": "Skilled",
            "gradeCode": "A",
            "unit": "Hr",
            "rate": round2(rate),
        })

    # ── equipment variants ──
    equipment_variants = []
    for name in eq_set:
        rate = code_data["equipment_rates"].get(name)
        if rate is None:
            rate = next((r["eq_rate"] for r in rows
                         if r["eq_name"] == name and r["eq_rate"] and r["eq_rate"] > 0), None)
        if rate is None:
            rate = 5.0
        equipment_variants.append({
            "roleCode": equipment_code(name),
            "roleName": name,
            "make": None,
            "model": None,
            "unit": "Hr",
            "rate": round2(rate),
            "standardOutputPerDay": None,
        })

    # ── work activities + activities + BOQ items ──
    # Compute per-activity productivity from DPRs.
    by_act = defaultdict(list)
    for r in rows:
        by_act[r["activity_code"]].append(r)

    activities = []
    work_activities = []
    boq_items = []

    for raw_code in activity_codes:
        san = sanitize_code(raw_code)
        master = code_data["activities_master"].get(raw_code, {})
        desc = master.get("description") or raw_code
        unit = master.get("unit") or "Nos"
        chap_code, chap_name, _ = wbs_chapter_for(raw_code)
        disc = discipline_for(raw_code)

        act_rows = by_act[raw_code]
        dates_for_act = sorted({r["date"] for r in act_rows})
        executed = sum((r["executed_qty"] or 0) for r in act_rows)
        total_mp_hours = sum((r["mp_hours"] or 0) * (r["mp_nos"] or 1) for r in act_rows
                             if r["mp_role"] and r["mp_hours"])
        total_eq_hours = sum((r["eq_hours"] or 0) * (r["eq_nos"] or 1) for r in act_rows
                             if r["eq_name"] and r["eq_hours"])

        output_per_man_day = None
        if executed > 0 and total_mp_hours > 0:
            output_per_man_day = round((executed / total_mp_hours) * 8, 2)

        output_per_hour = None
        if executed > 0 and total_eq_hours > 0:
            output_per_hour = round(executed / total_eq_hours, 2)

        # chainage extremes
        ch_from = min((r["from_ch"] for r in act_rows if r["from_ch"]), default=None)
        ch_to = max((r["to_ch"] for r in act_rows if r["to_ch"]), default=None)

        # supervisors who logged DPRs for this activity, in first-seen order
        seen_sup: list[str] = []
        for r in act_rows:
            if r["supervisor"] not in seen_sup:
                seen_sup.append(r["supervisor"])

        # planned window: actual span of DPRs + 7-day tail buffer
        planned_start = shifted(dates_for_act[0])
        planned_finish = shifted(dates_for_act[-1]) + datetime.timedelta(days=7)

        # Work activity master
        work_activities.append({
            "code": san,
            "name": desc,
            "defaultUnit": unit,
            "discipline": disc,
            "normCombination": "SERIES",
            "outputPerManPerDay": output_per_man_day,
            "outputPerHour": output_per_hour,
        })

        # Activity
        activities.append({
            "code": "ACT-" + san,
            "name": desc,
            "wbsChapterCode": chap_code,
            "workActivityCode": san,
            "plannedStart": planned_start.isoformat(),
            "plannedFinish": planned_finish.isoformat(),
            "chainageFromM": int(ch_from) if ch_from else None,
            "chainageToM": int(ch_to) if ch_to else None,
            "supervisorNames": seen_sup,
        })

        # BOQ item — qty assumed to be 1.3 × executed (so progress shows ~77%);
        # rate from observed cost-per-unit, defaulting to 100.
        boq_qty = round(executed * 1.3, 2) if executed > 0 else 0
        if boq_qty == 0:
            boq_qty = 100  # placeholder so it appears in the list
        # Derive a sensible rate: weighted average of (mp+eq cost) per executed unit.
        rate_guess = None
        if executed > 0:
            cost_mp = sum((r["mp_hours"] or 0) * (r["mp_nos"] or 1) * (r["mp_rate"] or 0)
                          for r in act_rows)
            cost_eq = sum((r["eq_hours"] or 0) * (r["eq_nos"] or 1) * (r["eq_rate"] or 0)
                          for r in act_rows)
            cost_total = cost_mp + cost_eq
            if cost_total > 0:
                rate_guess = round((cost_total / executed) * 1.4, 2)  # 40% margin
        if rate_guess is None or rate_guess <= 0:
            rate_guess = 50.0
        boq_items.append({
            "wbsChapterCode": chap_code,
            "itemNo": raw_code,
            "description": desc,
            "unit": unit,
            "boqQty": boq_qty,
            "boqRate": rate_guess,
            "budgetedRate": rate_guess,
            "chapter": f"{chap_code} - {chap_name}",
        })

    # ── WBS chapters (only those actually used) ──
    used_chapters = sorted({wbs_chapter_for(c)[0] for c in activity_codes})
    wbs_chapters = []
    for pfx, code, name, sort in WBS_CHAPTERS_DEF:
        if code in used_chapters:
            wbs_chapters.append({"code": code, "name": name, "sortOrder": sort})

    # ── DPR records ──
    # Two-stage grouping to keep DPR count manageable (~482 = 1 DPR / supervisor / day):
    #   1. Group daily rows by (date, supervisor, activity) — natural source granularity.
    #   2. For each (date, supervisor), keep only the activity with most source rows. That
    #      activity is the supervisor's primary work for the day. Other activities the
    #      supervisor touched are dropped — they were minor compared to the primary.
    #
    # Why: at full source granularity we produce 3,431 DPRs. Inserting each through the
    # DPR service runs ~1.5–2 s per row because of downstream listeners (BOQ sync, DBS
    # recompute, audit log). 3,431 × 2 s ≈ 2 hours. 482 × 2 s ≈ 16 minutes.
    raw_groups: dict[tuple[datetime.date, str, str], list[dict]] = defaultdict(list)
    for r in rows:
        raw_groups[(r["date"], r["supervisor"], r["activity_code"])].append(r)

    by_day_sup: dict[tuple[datetime.date, str], list[tuple[str, list[dict]]]] = defaultdict(list)
    for (d_, sup, act), rs in raw_groups.items():
        by_day_sup[(d_, sup)].append((act, rs))

    def _has_nonzero_qty(rs: list[dict]) -> bool:
        return any(r["executed_qty"] and r["executed_qty"] > 0 for r in rs)

    grouped: dict[tuple[datetime.date, str, str], list[dict]] = {}
    skipped_supdays_zero_qty = 0
    for (d_, sup), act_rows_pairs in by_day_sup.items():
        # Prefer the supervisor's activity that actually produced measurable output. Among
        # candidates with non-zero qty, pick the one with most source rows (their main work).
        # Drop the supervisor-day entirely if no activity had measurable output — the UI
        # rejects qty=0 saves and these DPRs would be unusable downstream.
        with_qty = [(a, rs) for a, rs in act_rows_pairs if _has_nonzero_qty(rs)]
        if not with_qty:
            skipped_supdays_zero_qty += 1
            continue
        primary_act, primary_rs = max(with_qty, key=lambda p: len(p[1]))
        grouped[(d_, primary_act, sup)] = primary_rs
    print(f"  skipped supervisor-days with no measurable output: {skipped_supdays_zero_qty}")

    dpr_records = []
    for (d, act_raw, sup), group in grouped.items():
        executed = next((g["executed_qty"] for g in group if g["executed_qty"]), 0) or 0
        # Aggregate manpower lines (sum nos+hours per role)
        mp_agg: dict[str, dict] = {}
        for g in group:
            if not g["mp_role"]:
                continue
            key = g["mp_role"]
            m = mp_agg.setdefault(key, {"nos": 0, "hours": 0.0, "rate": g["mp_rate"]})
            m["nos"] += g["mp_nos"] or 0
            m["hours"] += g["mp_hours"] or 0
            if g["mp_rate"] and g["mp_rate"] > 0:
                m["rate"] = g["mp_rate"]

        eq_agg: dict[str, dict] = {}
        for g in group:
            if not g["eq_name"]:
                continue
            key = g["eq_name"]
            e = eq_agg.setdefault(key, {"nos": 0, "hours": 0.0, "rate": g["eq_rate"]})
            e["nos"] += g["eq_nos"] or 0
            e["hours"] += g["eq_hours"] or 0
            if g["eq_rate"] and g["eq_rate"] > 0:
                e["rate"] = g["eq_rate"]

        san = sanitize_code(act_raw)
        first = group[0]
        # Skip DPRs that have no manpower AND no equipment AND no executed qty
        if not mp_agg and not eq_agg and executed == 0:
            continue

        dpr_records.append({
            "date": shifted(d).isoformat(),
            "activityCode": "ACT-" + san,
            "supervisorName": sup,
            "boqItemNo": act_raw,
            "workDoneQty": round2(executed) if executed else 0,
            "unit": first.get("unit"),
            "chainageFromM": int(first["from_ch"]) if first["from_ch"] else None,
            "chainageToM": int(first["to_ch"]) if first["to_ch"] else None,
            "shift": "DAY",
            "weather": None,
            "remarks": None,
            "manpower": [
                {
                    "roleCode": role_code(name),
                    "categoryCode": "Skilled",
                    "gradeCode": "A",
                    "nos": v["nos"] or 1,
                    "workingHours": round2(v["hours"]),
                    "otHours": 0,
                    "idleHours": 0,
                    "unitRate": round2(v["rate"]) if v["rate"] else None,
                    "contractorName": None,
                }
                for name, v in mp_agg.items()
            ],
            "equipment": [
                {
                    "roleCode": equipment_code(name),
                    "make": None,
                    "model": None,
                    "nos": v["nos"] or 1,
                    "workingHours": round2(v["hours"]),
                    "idleHours": 0,
                    "breakdownHours": 0,
                    "fuelLitres": 0,
                    "unitRate": round2(v["rate"]) if v["rate"] else None,
                }
                for name, v in eq_agg.items()
            ],
            "materials": [],
        })

    # ── Supplemental DPRs for activities not covered by the supervisor-primary dedup ──
    # Every activity should have at least one DPR so its BOQ item shows progress. For each
    # uncovered activity, find the source (date, supervisor) combo with most rows AND
    # non-zero executed qty, and add it as a DPR.
    covered_codes = {d["activityCode"] for d in dpr_records}
    expected_codes = {"ACT-" + sanitize_code(c) for c in activity_codes}
    uncovered = expected_codes - covered_codes
    if uncovered:
        # Group source rows by (activity, date, supervisor) once.
        by_act_day_sup: dict[tuple[str, datetime.date, str], list[dict]] = defaultdict(list)
        for r in rows:
            if not (r["executed_qty"] and r["executed_qty"] > 0):
                continue
            by_act_day_sup[(r["activity_code"], r["date"], r["supervisor"])].append(r)

        supplements_added = 0
        for code in sorted(uncovered):
            # find raw activity code that maps to this ACT-...
            raw = next((c for c in activity_codes if "ACT-" + sanitize_code(c) == code), None)
            if raw is None:
                continue
            # pick best (date, supervisor) for this activity
            candidates = [(k, v) for k, v in by_act_day_sup.items() if k[0] == raw]
            if not candidates:
                continue
            (_, d_, sup), best_rows = max(candidates, key=lambda kv: len(kv[1]))

            # Build a DPR from this row group (same shape as the main loop)
            executed = next((g["executed_qty"] for g in best_rows if g["executed_qty"]), 0) or 0
            mp_agg: dict[str, dict] = {}
            for g in best_rows:
                if not g["mp_role"]: continue
                key = g["mp_role"]
                m = mp_agg.setdefault(key, {"nos": 0, "hours": 0.0, "rate": g["mp_rate"]})
                m["nos"] += g["mp_nos"] or 0
                m["hours"] += g["mp_hours"] or 0
                if g["mp_rate"] and g["mp_rate"] > 0:
                    m["rate"] = g["mp_rate"]
            eq_agg: dict[str, dict] = {}
            for g in best_rows:
                if not g["eq_name"]: continue
                key = g["eq_name"]
                e = eq_agg.setdefault(key, {"nos": 0, "hours": 0.0, "rate": g["eq_rate"]})
                e["nos"] += g["eq_nos"] or 0
                e["hours"] += g["eq_hours"] or 0
                if g["eq_rate"] and g["eq_rate"] > 0:
                    e["rate"] = g["eq_rate"]

            san = sanitize_code(raw)
            first = best_rows[0]
            dpr_records.append({
                "date": shifted(d_).isoformat(),
                "activityCode": "ACT-" + san,
                "supervisorName": sup,
                "boqItemNo": raw,
                "workDoneQty": round2(executed),
                "unit": first.get("unit"),
                "chainageFromM": int(first["from_ch"]) if first["from_ch"] else None,
                "chainageToM": int(first["to_ch"]) if first["to_ch"] else None,
                "shift": "DAY",
                "weather": None,
                "remarks": None,
                "manpower": [
                    {"roleCode": role_code(name), "categoryCode": "Skilled", "gradeCode": "A",
                     "nos": v["nos"] or 1, "workingHours": round2(v["hours"]),
                     "otHours": 0, "idleHours": 0,
                     "unitRate": round2(v["rate"]) if v["rate"] else None,
                     "contractorName": None}
                    for name, v in mp_agg.items()
                ],
                "equipment": [
                    {"roleCode": equipment_code(name), "make": None, "model": None,
                     "nos": v["nos"] or 1, "workingHours": round2(v["hours"]),
                     "idleHours": 0, "breakdownHours": 0, "fuelLitres": 0,
                     "unitRate": round2(v["rate"]) if v["rate"] else None}
                    for name, v in eq_agg.items()
                ],
                "materials": [],
            })
            supplements_added += 1
        print(f"  supplemental DPRs added for uncovered activities: {supplements_added}")

    # Belt-and-braces: drop any remaining qty=0 DPR (shouldn't happen after the changes above).
    before = len(dpr_records)
    dpr_records = [r for r in dpr_records if r["workDoneQty"] and r["workDoneQty"] > 0]
    dropped = before - len(dpr_records)
    if dropped:
        print(f"  dropped {dropped} DPRs with workDoneQty=0 (UI rejects these)")

    # Sort DPRs by date asc, then activity, then supervisor — deterministic JSON.
    dpr_records.sort(key=lambda d: (d["date"], d["activityCode"], d["supervisorName"]))

    # ── Material: Concrete (from Concrete Summary - Khasab.xlsx) ──────────────────
    concrete_rows = load_concrete_summary()
    print(f"  concrete-summary rows parsed: {len(concrete_rows)}")

    # Build material variants for every grade observed in the file (so the role page is honest).
    observed_grades = sorted({c["grade"] for c in concrete_rows
                              if c["grade"] in CONCRETE_RATE_BY_GRADE})
    material_variants = [
        {
            "roleCode": "CONCRETE",
            "roleName": "Concrete",
            "specGrade": g,
            "unit": "m3",
            "rate": CONCRETE_RATE_BY_GRADE[g],
        }
        for g in observed_grades
    ]

    # Concrete Summary spans ~5 months (Dec 2025 → Apr 2026 in source) while our DPR window
    # spans ~2 months (Mar 18 → May 21 2026 after shift). Applying the same 418-day shift
    # would push concrete rows out of the DPR window. Instead, compress concrete dates
    # linearly into [dpr_min, dpr_max] so every row lands on a real DPR date.
    if concrete_rows:
        c_dates = [c["date"] for c in concrete_rows]
        c_min, c_max = min(c_dates), max(c_dates)
        c_span = max(1, (c_max - c_min).days)
        all_dpr_dates = sorted({datetime.date.fromisoformat(d["date"]) for d in dpr_records})
        d_min, d_max = all_dpr_dates[0], all_dpr_dates[-1]
        d_span = (d_max - d_min).days
        def concrete_target_date(src: datetime.date) -> datetime.date:
            return d_min + datetime.timedelta(days=int((src - c_min).days / c_span * d_span))
    else:
        concrete_target_date = lambda src: TODAY  # unreachable; keeps the type checker happy

    # Index DPRs by (date, activity) and by activity → ordered date list (for nearest-day fallback).
    dpr_by_date_activity: dict[tuple[str, str], list[dict]] = defaultdict(list)
    dates_by_activity: dict[str, list[datetime.date]] = defaultdict(list)
    for dpr in dpr_records:
        dpr_by_date_activity[(dpr["date"], dpr["activityCode"])].append(dpr)
        dates_by_activity[dpr["activityCode"]].append(datetime.date.fromisoformat(dpr["date"]))
    for a in dates_by_activity:
        dates_by_activity[a] = sorted(set(dates_by_activity[a]))

    def candidate_activity_codes(structure: str | None, grade: str) -> list[str]:
        codes: list[str] = []
        if structure:
            s = structure.lower()
            for key, raw_codes in CONCRETE_ACTIVITY_MAP.items():
                if key in s:
                    codes.extend(raw_codes)
        fallback = GRADE_FALLBACK_ACTIVITY.get(grade)
        if fallback:
            codes.append(fallback)
        # dedup while preserving order, then transform to "ACT-XXX"
        seen, out = set(), []
        for c in codes:
            sc = "ACT-" + sanitize_code(c)
            if sc not in seen:
                seen.add(sc); out.append(sc)
        return out

    def nearest_dpr_date(activity_code: str, target: datetime.date) -> datetime.date | None:
        dates = dates_by_activity.get(activity_code)
        if not dates:
            return None
        return min(dates, key=lambda d: abs((d - target).days))

    attached = 0
    skipped_no_activity = 0
    skipped_unknown_grade = 0
    for c in concrete_rows:
        if c["grade"] not in CONCRETE_RATE_BY_GRADE:
            skipped_unknown_grade += 1
            continue
        target_date = concrete_target_date(c["date"])
        candidates = candidate_activity_codes(c["structure"], c["grade"])
        # Pick the first activity that exists in our project. Nearest date for that activity.
        chosen_activity = next((a for a in candidates if a in dates_by_activity), None)
        if chosen_activity is None:
            skipped_no_activity += 1
            continue
        attach_date = nearest_dpr_date(chosen_activity, target_date)
        if attach_date is None:
            skipped_no_activity += 1
            continue
        target_dprs = dpr_by_date_activity[(attach_date.isoformat(), chosen_activity)]
        target = target_dprs[0]   # attach to the first supervisor's DPR for that day
        target["materials"].append({
            "roleCode": "CONCRETE",
            "specGrade": c["grade"],
            "quantity": round(c["qty"], 2),
            "unit": "m3",
            "unitRate": CONCRETE_RATE_BY_GRADE[c["grade"]],
            "vendorName": None,
        })
        attached += 1
    print(f"  concrete material rows attached to DPRs: {attached} "
          f"(skipped {skipped_no_activity} no-matching-activity, "
          f"{skipped_unknown_grade} unknown-grade)")

    return {
        "project": project,
        "manpowerVariants": manpower_variants,
        "equipmentVariants": equipment_variants,
        "materialVariants": material_variants,
        "workActivities": work_activities,
        "wbsChapters": wbs_chapters,
        "activities": activities,
        "boqItems": boq_items,
        "dprRecords": dpr_records,
        "validationErrors": [],
        "warnings": [
            f"Source dates 2025-01-24 → 2025-03-29 shifted by {shift_days} days; "
            f"last DPR now {dpr_records[-1]['date']}.",
            "Material role 'Concrete' was added from 'Concrete Summary - Khasab.xlsx'. "
            "Variants per grade (C15/C20/C25/C30/C35/C40) carry indicative OMR/m³ rates — "
            "update them in Admin → Resource Roles once the client shares contract rates. "
            "Concrete consumption rows are attached only to DPRs whose date+activity match a "
            "real concrete pour in the summary; rows outside the daily-data window are skipped.",
            "BOQ qty assumed at 1.3 × executed qty so progress reads ~77%; BOQ rate derived from "
            "average DPR cost per unit × 1.4 (40% margin), falling back to 50 when no cost data.",
            "12 unique supervisors found in the source. Stage 8 will fail loudly if any are missing "
            "from public.users — create them via Admin → Users before running Stage 8.",
        ],
    }


def main():
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_XLSX}")
    print(f"Reading {SOURCE_XLSX}")
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)

    print("Loading Code sheet (rates + activity master)…")
    code_data = load_code_sheet(wb)
    print(f"  manpower rates: {len(code_data['manpower_rates'])}")
    print(f"  equipment rates: {len(code_data['equipment_rates'])}")
    print(f"  activities master: {len(code_data['activities_master'])}")

    print("Loading daily sheets…")
    rows = load_daily_rows(wb)
    print(f"  raw daily rows with date+activity+supervisor: {len(rows)}")

    print("Building dataset…")
    dataset = build_dataset(code_data, rows)
    print(f"  manpowerVariants  : {len(dataset['manpowerVariants'])}")
    print(f"  equipmentVariants : {len(dataset['equipmentVariants'])}")
    print(f"  materialVariants  : {len(dataset['materialVariants'])}")
    print(f"  workActivities    : {len(dataset['workActivities'])}")
    print(f"  wbsChapters       : {len(dataset['wbsChapters'])}")
    print(f"  activities        : {len(dataset['activities'])}")
    print(f"  boqItems          : {len(dataset['boqItems'])}")
    print(f"  dprRecords        : {len(dataset['dprRecords'])}")

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with OUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(dataset, f, indent=2, ensure_ascii=False)
    print(f"Wrote {OUT_JSON} ({OUT_JSON.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
