#!/usr/bin/env python3
"""Parse Khasab daily-data workbook, shift dates +1y, group into DPR payloads.

Reads the 3 monthly sheets (Jan-2026, Feb-2026, March-2026), shifts dates from
2025 → 2026, groups child rows under one DPR header per (date, activity_code,
supervisor), and writes /tmp/khasab-dpr-parsed.json plus a validation report.
"""
import json
import sys
import re
import os
from collections import defaultdict
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

XLSX = os.environ.get("BIPROS_EXCEL_DIR", "./data/khasab-excel") + "/daily-data.xlsx"
OUT_JSON = os.environ.get("BIPROS_WORK_DIR", "/tmp/khasab") + "/khasab-dpr-parsed.json"
OUT_REPORT = os.environ.get("BIPROS_WORK_DIR", "/tmp/khasab") + "/khasab-dpr-validation_report.md"

# Canonical supervisor name → username
# User's required 8 supervisors + 4 extras found in actual data
# (Sohail, Manzar, V.P. Gupta, A.K. Mishra) so we don't drop ~2,659 DPRs.
SUPERVISOR_MAP = {
    "Mohd Ismaila":   "ismaila",
    "Md Saiffuddin":  "saiffuddin",
    "Illayaraja":     "illayaraja",
    "K. Barman":      "kbarman",
    "Vijaykumar":     "vijaykumar",
    "VijayKumar":     "vijaykumar",
    "Parvaiz":        "parvaiz",
    "Sanjar Alam":    "sanjar",
    "Anirban Datta":  "anirban",
    # Extras from actual data (Finding 10 in HTML log):
    "Sohail":         "sohail",
    "Manzar":         "manzar",
    "V.P. Gupta":     "vpgupta",
    "A.K.mishra":     "akmishra",
    "A.K. Mishra":    "akmishra",
}

SHEETS = ["Jan-2026", "Feb-2026", "March-2026"]

def shift_date(d):
    """Shift +1 year. Accept datetime/date/str; return ISO date string."""
    if d is None or d == "":
        return None
    if isinstance(d, str):
        try:
            d = datetime.fromisoformat(d.replace(" ", "T"))
        except Exception:
            return None
    if isinstance(d, datetime):
        d = d.date()
    if not isinstance(d, date):
        return None
    return (d + relativedelta(years=1)).isoformat()


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def safe_num(v):
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


wb = load_workbook(XLSX, data_only=True)
unknown_sups = defaultdict(int)
activity_codes = defaultdict(int)
activity_units = defaultdict(set)   # code -> {unit, ...}
manpower_roles = defaultdict(int)
equipment_names = defaultdict(int)
material_descs = defaultdict(int)
subcontractor_names = defaultdict(int)
total_rows = 0
skipped_rows = 0
skip_reasons = defaultdict(int)

groups = {}

for sheet in SHEETS:
    if sheet not in wb.sheetnames:
        print(f"WARN sheet missing: {sheet}", file=sys.stderr)
        continue
    ws = wb[sheet]
    # Header at row 4, data starts row 5 (per exploration)
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or len(row) < 11:
            continue
        raw_date = row[1]   # col B
        sup_name = safe_str(row[10])  # col K = Name
        act_code = safe_str(row[7])   # col H
        if not raw_date and not sup_name and not act_code:
            continue
        total_rows += 1

        date_iso = shift_date(raw_date)
        if not date_iso:
            skipped_rows += 1
            skip_reasons["bad_date"] += 1
            continue

        if not act_code:
            skipped_rows += 1
            skip_reasons["no_activity_code"] += 1
            continue
        activity_codes[act_code] += 1

        unit = safe_str(row[8])
        if unit:
            activity_units[act_code].add(unit)

        if sup_name and sup_name not in SUPERVISOR_MAP:
            unknown_sups[sup_name] += 1

        username = SUPERVISOR_MAP.get(sup_name)
        if not username:
            skipped_rows += 1
            skip_reasons["unknown_supervisor"] += 1
            continue

        site      = safe_str(row[2])
        location  = safe_str(row[3])
        loc_from  = safe_str(row[4])
        loc_to    = safe_str(row[5])
        side      = safe_str(row[6])
        qty_exec  = safe_num(row[9])

        mp_role   = safe_str(row[11])
        mp_count  = safe_num(row[12])
        mp_hours  = safe_num(row[13])
        mp_rate   = safe_num(row[14])
        mp_cost   = safe_num(row[15])

        eq_name   = safe_str(row[16])
        eq_count  = safe_num(row[17])
        eq_hours  = safe_num(row[18])
        eq_rate   = safe_num(row[19])
        eq_cost   = safe_num(row[20])

        mat_desc  = safe_str(row[21])
        mat_unit  = safe_str(row[22])
        mat_qty   = safe_num(row[23])
        mat_rate  = safe_num(row[24])
        mat_cost  = safe_num(row[25])

        sub_name  = safe_str(row[26])
        sub_desc  = safe_str(row[27])
        sub_unit  = safe_str(row[28])
        sub_qty   = safe_num(row[29])
        sub_rate  = safe_num(row[30])
        sub_cost  = safe_num(row[31]) if len(row) > 31 else 0

        if mp_role: manpower_roles[mp_role] += 1
        if eq_name: equipment_names[eq_name] += 1
        if mat_desc: material_descs[mat_desc] += 1
        if sub_name: subcontractor_names[sub_name] += 1

        key = (date_iso, act_code, username)
        if key not in groups:
            groups[key] = {
                "date": date_iso,
                "activity_code": act_code,
                "supervisor_username": username,
                "site": site,
                "location": location,
                "loc_from": loc_from,
                "loc_to": loc_to,
                "side": side,
                "unit": unit,
                "qty_executed": qty_exec,
                "manpower": [],
                "equipment": [],
                "material": [],
                "subcontractor": [],
            }
        else:
            # If a later row has a non-zero qty_executed, keep the max
            # (sometimes the qty appears on only the first row of a group)
            if qty_exec and qty_exec > groups[key].get("qty_executed", 0):
                groups[key]["qty_executed"] = qty_exec
            if unit and not groups[key].get("unit"):
                groups[key]["unit"] = unit

        g = groups[key]
        # A blank "Nr." column means 1 — it is blank on ~94% of manpower and ~92% of
        # equipment rows in the source workbook. Gate inclusion on the role/name (plus any
        # of count/hours/rate) so those rows aren't dropped; default the count to 1.
        if mp_role and (mp_count or mp_hours or mp_rate):
            g["manpower"].append({
                "role": mp_role, "count": int(mp_count) if mp_count else 1,
                "hours": mp_hours, "rate": mp_rate, "cost": mp_cost
            })
        if eq_name and (eq_count or eq_hours or eq_rate):
            g["equipment"].append({
                "name": eq_name, "count": int(eq_count) if eq_count else 1,
                "hours": eq_hours, "rate": eq_rate, "cost": eq_cost
            })
        if mat_desc and mat_qty:
            g["material"].append({
                "desc": mat_desc, "unit": mat_unit,
                "qty": mat_qty, "rate": mat_rate, "cost": mat_cost
            })
        if sub_name and sub_qty:
            g["subcontractor"].append({
                "name": sub_name, "desc": sub_desc, "unit": sub_unit,
                "qty": sub_qty, "rate": sub_rate, "cost": sub_cost
            })

dprs = list(groups.values())

with open(OUT_JSON, "w") as f:
    json.dump(dprs, f, indent=2, default=str)

# DPR counts per month
by_month = defaultdict(int)
for d in dprs:
    by_month[d["date"][:7]] += 1

# Write report
with open(OUT_REPORT, "w") as f:
    f.write(f"# Khasab DPR Parse — Validation Report\n\n")
    f.write(f"Source: `{XLSX}`\n\n")
    f.write(f"## Summary\n\n")
    f.write(f"| Metric | Value |\n|---|---:|\n")
    f.write(f"| Total source rows scanned | {total_rows} |\n")
    f.write(f"| Source rows skipped | {skipped_rows} |\n")
    f.write(f"| Unique DPR groups produced | {len(dprs)} |\n")
    f.write(f"| Collapse ratio | {total_rows/max(len(dprs),1):.1f}x |\n\n")
    f.write(f"## Skip reasons\n\n")
    for reason, n in sorted(skip_reasons.items(), key=lambda x:-x[1]):
        f.write(f"- `{reason}`: {n}\n")
    f.write(f"\n## DPRs per month (after +1y shift)\n\n")
    for m in sorted(by_month):
        f.write(f"- {m}: {by_month[m]}\n")
    f.write(f"\n## Unknown supervisors ({len(unknown_sups)})\n\n")
    if unknown_sups:
        f.write("| Source name | Row count |\n|---|---:|\n")
        for n, c in sorted(unknown_sups.items(), key=lambda x:-x[1]):
            f.write(f"| `{n}` | {c} |\n")
    else:
        f.write("(none)\n")
    f.write(f"\n## Activity codes ({len(activity_codes)})\n\n")
    f.write("| Code | Row count | Unit(s) |\n|---|---:|---|\n")
    for code, c in sorted(activity_codes.items(), key=lambda x:-x[1]):
        units = ",".join(sorted(activity_units[code]))
        f.write(f"| `{code}` | {c} | {units} |\n")
    f.write(f"\n## Manpower roles ({len(manpower_roles)})\n\n")
    f.write("| Role | Occurrences |\n|---|---:|\n")
    for r, c in sorted(manpower_roles.items(), key=lambda x:-x[1]):
        f.write(f"| `{r}` | {c} |\n")
    f.write(f"\n## Equipment names ({len(equipment_names)})\n\n")
    f.write("| Equipment | Occurrences |\n|---|---:|\n")
    for r, c in sorted(equipment_names.items(), key=lambda x:-x[1])[:30]:
        f.write(f"| `{r}` | {c} |\n")
    f.write(f"\n## Material descriptions ({len(material_descs)})\n\n")
    f.write("| Material | Occurrences |\n|---|---:|\n")
    for r, c in sorted(material_descs.items(), key=lambda x:-x[1])[:30]:
        f.write(f"| `{r}` | {c} |\n")
    f.write(f"\n## Subcontractors ({len(subcontractor_names)})\n\n")
    f.write("| Subcontractor | Occurrences |\n|---|---:|\n")
    for r, c in sorted(subcontractor_names.items(), key=lambda x:-x[1])[:30]:
        f.write(f"| `{r}` | {c} |\n")

print(f"Wrote {OUT_JSON} ({len(dprs)} DPRs)")
print(f"Wrote {OUT_REPORT}")
print(f"Total rows scanned: {total_rows}, skipped: {skipped_rows}")
print(f"DPRs/month: {dict(by_month)}")
print(f"Unknown supervisors: {len(unknown_sups)}")
