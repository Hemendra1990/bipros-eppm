#!/usr/bin/env python3
"""Extract activity master from DBS workbook 'DPR' sheet.
Produces /tmp/khasab/activity-master.json: {code: {name, unit, unit_rate, plan_qty}}
"""
import json
import os
import re
from openpyxl import load_workbook

XLSX = "/Volumes/Java/Projects/bipros-eppm/docs/ActualData/3. Supervisor-Engineer-CM-PM DBS (2).xlsx"
OUT_MASTER = "/tmp/khasab/activity-master.json"
os.makedirs(os.path.dirname(OUT_MASTER), exist_ok=True)

wb = load_workbook(XLSX, data_only=True)
ws = wb["DPR"]

master = {}
current_chapter = ""

for r in range(5, ws.max_row + 1):
    code_col = ws.cell(row=r, column=2).value  # 'Code' (chapter)
    boq_col = ws.cell(row=r, column=3).value   # 'BOQ #' (real activity code)
    desc = ws.cell(row=r, column=4).value      # 'Description'
    unit = ws.cell(row=r, column=5).value
    rate = ws.cell(row=r, column=6).value
    qty  = ws.cell(row=r, column=7).value

    if not boq_col or not desc:
        continue

    desc_str = str(desc).strip()
    code_key = str(boq_col).strip()

    # Skip the trailing test/junk row where desc is just a digit or "Total Amount"
    if desc_str in ("1", "2", "3", "4", "Total Amount") or len(desc_str) <= 2:
        continue
    # Skip chapter-summary rows where code_col=boq_col=desc is the chapter title
    if code_col and code_key == str(code_col).strip():
        # This is the chapter header — name is the chapter name (e.g. "Preliminaries")
        # Keep it but tagged as chapter
        master[code_key] = {
            "code": code_key,
            "name": desc_str,
            "unit": (str(unit).strip() if unit else "LS"),
            "unit_rate": float(rate) if rate else 0,
            "plan_qty": float(qty) if qty else 0,
            "chapter": desc_str,
            "is_chapter": True,
        }
        current_chapter = desc_str
        continue

    master[code_key] = {
        "code": code_key,
        "name": desc_str,
        "unit": (str(unit).strip() if unit else "nos"),
        "unit_rate": float(rate) if rate else 0,
        "plan_qty": float(qty) if qty else 0,
        "chapter": current_chapter,
        "is_chapter": False,
    }

with open(OUT_MASTER, "w") as f:
    json.dump(master, f, indent=2)

print(f"Extracted {len(master)} activity master rows")
print(f"\nSample entries:")
for k in list(master.keys())[:8]:
    m = master[k]
    print(f"  {k}: {m['name'][:60]}  [{m['unit']}]  chapter='{m['chapter'][:30]}'")
print(f"\nWrote {OUT_MASTER}")

def normalize_code(c):
    """Normalize spacing in activity codes: '2.1.5 (i)' → '2.1.5(i)'"""
    return re.sub(r'\s+', '', str(c)).strip()

# Re-key master by normalized code
master_norm = {normalize_code(k): v for k, v in master.items()}

# Cross-check against Khasab DPR data using normalized codes
DPRS = json.load(open("/tmp/khasab-dpr-parsed.json"))
khasab_codes_raw = set(d["activity_code"] for d in DPRS)
khasab_codes_norm = {normalize_code(c): c for c in khasab_codes_raw}

matched = set(khasab_codes_norm.keys()) & set(master_norm.keys())
missing = set(khasab_codes_norm.keys()) - set(master_norm.keys())
print(f"\n=== Khasab vs Master cross-check ===")
print(f"Khasab activity codes: {len(khasab_codes_raw)}")
print(f"Matched in master: {len(matched)}")
print(f"Missing from master: {len(missing)}")
if missing:
    print(f"Missing normalized codes (need fallback names):")
    for c in sorted(missing):
        print(f"  - '{c}' (was: '{khasab_codes_norm[c]}')")

# Save the normalized lookup
with open("/tmp/khasab/activity-master-normalized.json", "w") as f:
    json.dump(master_norm, f, indent=2)
print(f"\nWrote /tmp/khasab/activity-master-normalized.json (keyed by normalized code)")
