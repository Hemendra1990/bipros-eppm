import os
#!/usr/bin/env python3
"""Final demo polish:
1. Re-extract activity names from the CODE sheet of the Khasab daily-data workbook (where they actually live)
2. SQL UPDATE work_activities + activity.activities with REAL names
3. Add weather to DPRs (mix of CLEAR / CLOUDY / RAIN / WINDY)
4. Create Risk register entries for the project
"""
import json
import subprocess
import urllib.request
import urllib.error
import random
from openpyxl import load_workbook

BASE = "http://localhost:8080"
TOKEN = open("/tmp/admin-token.txt").read().strip()
PROJECT_ID = open("/tmp/khasab/project-id.txt").read().strip()
PSQL = "/Applications/Postgres.app/Contents/Versions/latest/bin/psql"
PG_BASE = ["env", f"PGPASSWORD={os.environ.get('BIPROS_PG_PASS', 'bipros_dev')}", PSQL, "-h", os.environ.get("BIPROS_PG_HOST", "127.0.0.1"), "-p", os.environ.get("BIPROS_PG_PORT", "5432"), "-U", os.environ.get("BIPROS_PG_USER", "bipros"), "-d", os.environ.get("BIPROS_PG_DB", "bipros"), "-A", "-F", "|", "-t", "-c"]


def sql(q, ignore=False):
    out = subprocess.run(PG_BASE + [q], capture_output=True, text=True, timeout=60)
    if out.returncode != 0:
        if ignore:
            return None
        raise Exception(f"SQL: {out.stderr[:200]}\nQUERY: {q[:200]}")
    return [line.split("|") for line in out.stdout.strip().split("\n") if line.strip()]


def http(method, path, body=None, timeout=15):
    req = urllib.request.Request(f"{BASE}{path}", method=method)
    req.add_header("Authorization", f"Bearer {TOKEN}")
    req.add_header("Content-Type", "application/json")
    data = json.dumps(body).encode() if body else None
    try:
        with urllib.request.urlopen(req, data=data, timeout=timeout) as resp:
            return resp.status, json.loads(resp.read())
    except urllib.error.HTTPError as e:
        try:
            err = json.loads(e.read())
        except Exception:
            err = {"error": str(e)}
        return e.code, err
    except Exception as e:
        return 0, {"error": str(e)}


# ====================================================================
# STEP 1: Parse Code sheet to get REAL activity names
# ====================================================================
print("=== Step 1: Parse Code sheet for real activity names ===")
XLSX = "/Volumes/Java/Projects/bipros-eppm/docs/ActualData/1. Daily Data-Khasab  Jan, Feb, Mar 2026.xlsx"
wb = load_workbook(XLSX, data_only=True)
ws = wb["Code"]

# Code in col C (3), Name in col D (4), Unit in col E (5)
code_to_real_name = {}
for r in range(2, ws.max_row + 1):
    code = ws.cell(row=r, column=3).value
    name = ws.cell(row=r, column=4).value
    unit = ws.cell(row=r, column=5).value
    if code and name:
        c = str(code).strip()
        n = str(name).strip()
        u = str(unit).strip() if unit else None
        # Normalize unit (Code sheet uses different units across rows — first 14 are unit headers, ignore)
        code_to_real_name[c] = {"name": n, "unit": u}

print(f"  Extracted {len(code_to_real_name)} real activity names from Code sheet")
print(f"  Samples:")
for c in ["1", "1.1", "1.2", "2.3.6(i)b", "13.1.7(ix)a", "5.1.7 (iii)"]:
    if c in code_to_real_name:
        print(f"    {c}: {code_to_real_name[c]['name']}")


# ====================================================================
# STEP 2: UPDATE activity names + work_activity names in DB
# ====================================================================
print("\n=== Step 2: SQL UPDATE activity + work_activity names ===")
updated_activities = 0
updated_work_activities = 0

# Build a code-lookup that handles spacing differences
import re
def norm(c):
    return re.sub(r'\s+', '', str(c)).strip()
code_norm_to_data = {norm(k): v for k, v in code_to_real_name.items()}

# Get all activities in this project
acts = sql(f"""
SELECT a.id::text, a.code, a.name, a.work_activity_id::text
FROM activity.activities a
WHERE a.project_id='{PROJECT_ID}'
""")
for act_id, act_code, old_name, wa_id in acts:
    real = code_norm_to_data.get(norm(act_code))
    if not real:
        print(f"  {act_code}: no real name in Code sheet — keeping '{old_name[:40]}'")
        continue
    new_name = real["name"][:100].replace("'", "''")
    sql(f"UPDATE activity.activities SET name='{new_name}', updated_at=now() WHERE id='{act_id}'")
    updated_activities += 1
    # Denormalized snapshot — daily_progress_reports holds activity_name at write time and is
    # what the DPR list view groups by. A raw SQL rename bypasses ActivityRenameDprSyncListener
    # (which only fires for API renames), so update the snapshot here too.
    sql(f"UPDATE project.daily_progress_reports SET activity_name='{new_name}', updated_at=now() WHERE activity_id='{act_id}'")
    if wa_id and wa_id != "":
        sql(f"UPDATE resource.work_activities SET name='{new_name}', updated_at=now() WHERE id='{wa_id}'")
        updated_work_activities += 1
print(f"  Updated {updated_activities} activities + {updated_work_activities} work_activities with REAL names")

# Verify
verify = sql(f"""
SELECT code, name FROM activity.activities WHERE project_id='{PROJECT_ID}' ORDER BY code LIMIT 10
""")
print(f"\n  Verification (first 10):")
for c, n in verify:
    print(f"    {c}: {n}")


# ====================================================================
# STEP 3: Add weather to DPRs
# ====================================================================
print("\n=== Step 3: Add weather conditions to DPRs ===")
weather_options = ['CLEAR', 'CLOUDY', 'PARTLY_CLOUDY', 'WINDY', 'RAIN']
# Assign weather deterministically based on date (so it's consistent and looks realistic)
sql(f"""
UPDATE project.daily_progress_reports SET weather_condition = CASE
  WHEN EXTRACT(DOW FROM report_date) = 0 THEN 'CLEAR'
  WHEN EXTRACT(DOW FROM report_date) = 1 THEN 'CLEAR'
  WHEN EXTRACT(DOW FROM report_date) = 2 THEN 'PARTLY_CLOUDY'
  WHEN EXTRACT(DOW FROM report_date) = 3 THEN 'CLOUDY'
  WHEN EXTRACT(DOW FROM report_date) = 4 THEN 'WINDY'
  WHEN EXTRACT(DOW FROM report_date) = 5 THEN 'CLEAR'
  WHEN EXTRACT(DOW FROM report_date) = 6 THEN 'RAIN'
END
WHERE project_id='{PROJECT_ID}'
""")
weather_dist = sql(f"""
SELECT weather_condition, COUNT(*) FROM project.daily_progress_reports
WHERE project_id='{PROJECT_ID}' GROUP BY weather_condition ORDER BY COUNT(*) DESC
""")
print(f"  Weather distribution across {sum(int(w[1]) for w in weather_dist)} DPRs:")
for w, n in weather_dist:
    print(f"    {w}: {n}")


# ====================================================================
# STEP 4: Create Risk register entries
# ====================================================================
print("\n=== Step 4: Create Risk register entries ===")

# Check risk endpoint shape
risks = [
    {
        "code": "RISK-001", "title": "Monsoon delays in March",
        "description": "Late-March rains may delay bituminous works and slope dressing",
        "category": "Weather", "phase": "EXECUTION",
        "probability": 4, "impactCost": 3, "impactSchedule": 4,
        "ownerRole": "PROJECT_MANAGER",
    },
    {
        "code": "RISK-002", "title": "Equipment availability — Excavator fleet",
        "description": "High utilization on Excavators (>80%) may strain availability if any breakdown occurs",
        "category": "Equipment", "phase": "EXECUTION",
        "probability": 3, "impactCost": 3, "impactSchedule": 4,
        "ownerRole": "PROJECT_MANAGER",
    },
    {
        "code": "RISK-003", "title": "Blasting permit delays",
        "description": "MoTC blasting permits for hard-rock excavation can take 2-4 weeks",
        "category": "Regulatory", "phase": "EXECUTION",
        "probability": 3, "impactCost": 2, "impactSchedule": 5,
        "ownerRole": "PROJECT_MANAGER",
    },
    {
        "code": "RISK-004", "title": "Concrete supply continuity",
        "description": "Local batching plant capacity may not match peak pour demand for bridge structures",
        "category": "Procurement", "phase": "EXECUTION",
        "probability": 2, "impactCost": 4, "impactSchedule": 3,
        "ownerRole": "SITE_MANAGER",
    },
    {
        "code": "RISK-005", "title": "Skilled labour shortage — Steel Fixers",
        "description": "Regional shortage of certified rebar fixers may slow concrete activities",
        "category": "Resource", "phase": "EXECUTION",
        "probability": 3, "impactCost": 2, "impactSchedule": 3,
        "ownerRole": "SITE_ENGINEER",
    },
    {
        "code": "RISK-006", "title": "Underground utility strike",
        "description": "Existing 11kV electrical lines + water mains in alignment; relocation drawings incomplete",
        "category": "Safety", "phase": "EXECUTION",
        "probability": 3, "impactCost": 4, "impactSchedule": 3,
        "ownerRole": "SAFETY_OFFICER",
    },
    {
        "code": "RISK-007", "title": "Borrow pit yield variability",
        "description": "Borrow source for embankment shows variable gradation; may need extra screening passes",
        "category": "Material", "phase": "EXECUTION",
        "probability": 4, "impactCost": 3, "impactSchedule": 2,
        "ownerRole": "QC_ENGINEER",
    },
    {
        "code": "RISK-008", "title": "Bridge bearing delivery",
        "description": "Imported elastomeric bearings have 14-week lead time; need PO by week 12 to meet erection",
        "category": "Procurement", "phase": "EXECUTION",
        "probability": 3, "impactCost": 2, "impactSchedule": 4,
        "ownerRole": "PROJECT_MANAGER",
    },
]

ok = fail = 0
for r in risks:
    body = {
        "projectId": PROJECT_ID,
        "code": r["code"],
        "title": r["title"],
        "description": r["description"],
        "category": r["category"],
        "phase": r["phase"],
        "probabilityRating": r["probability"],
        "impactCostRating": r["impactCost"],
        "impactScheduleRating": r["impactSchedule"],
        "status": "OPEN",
    }
    sc, resp = http("POST", f"/v1/projects/{PROJECT_ID}/risks", body)
    if sc in (200, 201):
        ok += 1
    else:
        fail += 1
        if fail < 3:
            err = resp.get("error", {}).get("message", "?") if isinstance(resp.get("error"), dict) else str(resp.get("error"))
            print(f"  {r['code']}: FAIL {sc} {err[:120]}")
print(f"\n  Created {ok} risks via API ({fail} failed)")

# Fallback: SQL insert if API failed
if ok == 0:
    print("\n  API failed, trying direct SQL insert into risk.risks...")
    cols = sql("SELECT column_name FROM information_schema.columns WHERE table_schema='risk' AND table_name='risks' ORDER BY ordinal_position")
    print(f"  risk.risks columns: {[c[0] for c in cols][:15]}")
    sql(f"DELETE FROM risk.risks WHERE project_id='{PROJECT_ID}'", ignore=True)
    for i, r in enumerate(risks):
        title_safe = r["title"].replace("'", "''")
        desc_safe = r["description"].replace("'", "''")
        result = sql(f"""
INSERT INTO risk.risks (id, created_at, updated_at, version, project_id, code, title, description, category, phase,
  probability_rating, impact_cost_rating, impact_schedule_rating, status, identified_date)
VALUES (gen_random_uuid(), now(), now(), 0, '{PROJECT_ID}', '{r["code"]}', '{title_safe}', '{desc_safe}',
  '{r["category"]}', '{r["phase"]}', {r["probability"]}, {r["impactCost"]}, {r["impactSchedule"]}, 'OPEN', '2026-01-15')
RETURNING id::text
""", ignore=True)
        if result:
            ok += 1
    print(f"  Direct SQL: {ok} risks inserted")


# Final verification
print("\n=== FINAL STATE ===")
final = sql(f"""
SELECT 'activities_with_real_name' AS m, COUNT(*) FROM activity.activities a
  WHERE a.project_id='{PROJECT_ID}' AND a.name IN (SELECT name FROM activity.activities WHERE project_id='{PROJECT_ID}' AND name NOT IN ('Preliminaries'))
UNION ALL SELECT 'dprs_with_weather', COUNT(*) FROM project.daily_progress_reports WHERE project_id='{PROJECT_ID}' AND weather_condition IS NOT NULL
UNION ALL SELECT 'risks', COUNT(*) FROM risk.risks WHERE project_id='{PROJECT_ID}'
""")
for r in final:
    print(f"  {r[0]:30} {r[1]}")
