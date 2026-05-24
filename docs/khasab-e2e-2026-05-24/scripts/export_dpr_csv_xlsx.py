#!/usr/bin/env python3
"""Export Khasab DPRs to flat CSV + 4-sheet XLSX."""
import csv
import os
import subprocess
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

REPO = "/Volumes/Java/Projects/bipros-eppm"
PROJECT_ID = open("/tmp/khasab/project-id.txt").read().strip()
PSQL = "/Applications/Postgres.app/Contents/Versions/latest/bin/psql"
PG_BASE = ["env", "PGPASSWORD=bipros_dev", PSQL, "-h", "127.0.0.1", "-U", "bipros", "-d", "bipros", "-A", "-F", ",", "-t", "-c"]
OUT_DIR = f"{REPO}/docs/ActualData/exports"
os.makedirs(OUT_DIR, exist_ok=True)


def csv_rows(sql):
    full = sql.replace("$PROJECT_ID", PROJECT_ID)
    out = subprocess.run(PG_BASE + [full], capture_output=True, text=True, timeout=60)
    if out.returncode != 0:
        print(f"SQL error: {out.stderr}")
        return []
    return [line.split(",") for line in out.stdout.strip().split("\n") if line.strip()]


# Main flat CSV
SQL = """
SELECT
  d.report_date,
  u.username AS supervisor,
  a.code AS activity_code,
  w.code AS wbs_code,
  w.name AS wbs_name,
  d.qty_executed,
  d.unit,
  COALESCE((SELECT SUM(line_cost) FROM project.dpr_manpower m WHERE m.dpr_id=d.id), 0) AS manpower_cost,
  COALESCE((SELECT SUM(line_cost) FROM project.dpr_equipment e WHERE e.dpr_id=d.id), 0) AS equipment_cost,
  COALESCE((SELECT COUNT(*) FROM project.dpr_manpower m WHERE m.dpr_id=d.id), 0) AS manpower_lines,
  COALESCE((SELECT COUNT(*) FROM project.dpr_equipment e WHERE e.dpr_id=d.id), 0) AS equipment_lines,
  d.remarks
FROM project.daily_progress_reports d
LEFT JOIN public.users u ON u.id = d.supervisor_user_id
LEFT JOIN activity.activities a ON a.id = d.activity_id
LEFT JOIN project.wbs_nodes w ON w.id = a.wbs_node_id
WHERE d.project_id = '$PROJECT_ID'
ORDER BY d.report_date, a.code, u.username
"""

rows = csv_rows(SQL)
print(f"Exporting {len(rows)} DPRs...")

# CSV
csv_path = f"{OUT_DIR}/khasab-dpr-2026-05-24.csv"
HEADERS = ["report_date", "supervisor", "activity_code", "wbs_code", "wbs_name",
           "qty_executed", "unit", "manpower_cost", "equipment_cost",
           "manpower_lines", "equipment_lines", "remarks"]
with open(csv_path, "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(HEADERS)
    w.writerows(rows)
print(f"  → {csv_path}")

# XLSX — 4 sheets
xlsx_path = f"{OUT_DIR}/khasab-dpr-2026-05-24.xlsx"
wb = Workbook()
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1f2937")

# Sheet 1: DPR (flat)
ws = wb.active
ws.title = "DPR"
ws.append(HEADERS)
for i, h in enumerate(HEADERS, 1):
    ws.cell(row=1, column=i).font = header_font
    ws.cell(row=1, column=i).fill = header_fill
for r in rows:
    ws.append(r)
for col in ws.columns:
    max_len = max(len(str(c.value)) if c.value else 0 for c in col[:20])
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

# Sheet 2: By Supervisor
by_sup = defaultdict(lambda: {"dprs": 0, "qty": 0.0, "mp_cost": 0.0, "eq_cost": 0.0})
for r in rows:
    by_sup[r[1]]["dprs"] += 1
    try:
        by_sup[r[1]]["qty"] += float(r[5]) if r[5] else 0
        by_sup[r[1]]["mp_cost"] += float(r[7]) if r[7] else 0
        by_sup[r[1]]["eq_cost"] += float(r[8]) if r[8] else 0
    except (ValueError, IndexError):
        pass
ws2 = wb.create_sheet("By-Supervisor")
ws2.append(["supervisor", "dpr_count", "total_qty", "total_manpower_cost", "total_equipment_cost"])
for i in range(1, 6):
    ws2.cell(row=1, column=i).font = header_font
    ws2.cell(row=1, column=i).fill = header_fill
for sup, s in sorted(by_sup.items(), key=lambda x: -x[1]["dprs"]):
    ws2.append([sup, s["dprs"], round(s["qty"], 2), round(s["mp_cost"], 2), round(s["eq_cost"], 2)])

# Sheet 3: By Activity
by_act = defaultdict(lambda: {"dprs": 0, "qty": 0.0, "mp_cost": 0.0, "eq_cost": 0.0})
for r in rows:
    by_act[r[2]]["dprs"] += 1
    try:
        by_act[r[2]]["qty"] += float(r[5]) if r[5] else 0
        by_act[r[2]]["mp_cost"] += float(r[7]) if r[7] else 0
        by_act[r[2]]["eq_cost"] += float(r[8]) if r[8] else 0
    except (ValueError, IndexError):
        pass
ws3 = wb.create_sheet("By-Activity")
ws3.append(["activity_code", "dpr_count", "total_qty", "total_manpower_cost", "total_equipment_cost"])
for i in range(1, 6):
    ws3.cell(row=1, column=i).font = header_font
    ws3.cell(row=1, column=i).fill = header_fill
for act, s in sorted(by_act.items(), key=lambda x: -x[1]["dprs"]):
    ws3.append([act, s["dprs"], round(s["qty"], 2), round(s["mp_cost"], 2), round(s["eq_cost"], 2)])

# Sheet 4: By Date
by_date = defaultdict(lambda: {"dprs": 0, "qty": 0.0, "mp_cost": 0.0, "eq_cost": 0.0})
for r in rows:
    by_date[r[0]]["dprs"] += 1
    try:
        by_date[r[0]]["qty"] += float(r[5]) if r[5] else 0
        by_date[r[0]]["mp_cost"] += float(r[7]) if r[7] else 0
        by_date[r[0]]["eq_cost"] += float(r[8]) if r[8] else 0
    except (ValueError, IndexError):
        pass
ws4 = wb.create_sheet("By-Date")
ws4.append(["date", "dpr_count", "total_qty", "total_manpower_cost", "total_equipment_cost"])
for i in range(1, 6):
    ws4.cell(row=1, column=i).font = header_font
    ws4.cell(row=1, column=i).fill = header_fill
for date, s in sorted(by_date.items()):
    ws4.append([date, s["dprs"], round(s["qty"], 2), round(s["mp_cost"], 2), round(s["eq_cost"], 2)])

wb.save(xlsx_path)
print(f"  → {xlsx_path}")
print(f"\nTotal DPRs exported: {len(rows)}")
