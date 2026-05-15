#!/usr/bin/env python3
"""Append many rows to docs/ai-test-results-2026-05-15.xlsx.

Reads a JSON array of row objects from stdin.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

OUT = Path("/Volumes/Java/Projects/bipros-eppm/docs/ai-test-results-2026-05-15.xlsx")
HEADERS = [
    "#",
    "Question",
    "Answer",
    "Tools Called",
    "Verified",
    "Verifying Seen",
    "Rounds",
    "Duration (ms)",
    "Error",
]


def main() -> None:
    raw = sys.stdin.read()
    rows = json.loads(raw)
    if isinstance(rows, str):
        rows = json.loads(rows)
    # Skip stop sentinels
    rows = [r for r in rows if "question" in r]
    OUT.parent.mkdir(parents=True, exist_ok=True)
    if OUT.exists():
        wb = load_workbook(OUT)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "AI Test"
        ws.append(HEADERS)
        for col, width in zip("ABCDEFGHI", [5, 60, 80, 40, 10, 14, 8, 14, 30]):
            ws.column_dimensions[col].width = width

    for row in rows:
        tools = row.get("tools") or []
        if isinstance(tools, list):
            tools = ", ".join(tools)
        ws.append([
            row.get("n"),
            row.get("question", ""),
            row.get("answer", ""),
            tools,
            str(row.get("verified")) if row.get("verified") is not None else "",
            "true" if row.get("verifying_seen") else "false",
            row.get("rounds"),
            row.get("duration_ms"),
            row.get("error") or "",
        ])
    wb.save(OUT)
    print(f"ok {len(rows)} rows", flush=True)


if __name__ == "__main__":
    main()
