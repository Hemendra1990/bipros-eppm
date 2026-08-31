#!/usr/bin/env python3
"""Append one row to docs/ai-test-results-2026-05-15.xlsx.

Reads a single JSON object from stdin with keys:
  n, question, answer, tools, verified, verifying_seen, rounds, duration_ms, error

Creates the workbook + header row on first use.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

OUT = Path("/Volumes/Java/Projects/bipros-eppm/docs/ai-test-results-2026-05-15.xlsx")
HEADERS = [
    "#",
    "Question",
    "Answer",
    "Tools Called",
    "Verified",
    "Verifying Seen",
    "Rounds",
    "Duration (ms)",
    "Error",
]


def main() -> None:
    row = json.loads(sys.stdin.read())
    OUT.parent.mkdir(parents=True, exist_ok=True)
    if OUT.exists():
        wb = load_workbook(OUT)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "AI Test"
        ws.append(HEADERS)
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 30

    ws.append([
        row.get("n"),
        row.get("question", ""),
        row.get("answer", ""),
        ", ".join(row.get("tools") or []) if isinstance(row.get("tools"), list) else (row.get("tools") or ""),
        str(row.get("verified")) if row.get("verified") is not None else "",
        "true" if row.get("verifying_seen") else "false",
        row.get("rounds"),
        row.get("duration_ms"),
        row.get("error") or "",
    ])
    wb.save(OUT)
    print(f"ok row #{row.get('n')}", flush=True)


if __name__ == "__main__":
    main()
